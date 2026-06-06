"""VSI tool input support app.

用途:
- 国別仕様組み合わせファイルから大分類と細目を抽出
- 車両情報一覧の細目列に不足している候補を抽出
- 記載方法シートの基準をもとに、LLMで追記要否を判定して反映

実行:
  streamlit run VSI_tool_input.py
"""

from __future__ import annotations

from copy import copy
import logging
import os
import re
from pathlib import Path
import time
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Dict, List, Optional, Set, Tuple

import streamlit as st
from openai import AzureOpenAI
from openpyxl import load_workbook
from openpyxl.styles import Color
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel


LOG_DIR = Path(__file__).resolve().parent / "logs"
EXEC_LOG_FILE = LOG_DIR / "vsi_tool_execution.log"
ERROR_LOG_FILE = LOG_DIR / "vsi_tool_error.log"


def setup_logger() -> logging.Logger:
	"""実行ログとエラーログをファイル出力するロガーを構成する。"""
	LOG_DIR.mkdir(parents=True, exist_ok=True)
	logger = logging.getLogger("vsi_tool_input")
	logger.setLevel(logging.INFO)

	if logger.handlers:
		return logger

	formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

	exec_handler = logging.FileHandler(EXEC_LOG_FILE, encoding="utf-8")
	exec_handler.setLevel(logging.INFO)
	exec_handler.setFormatter(formatter)

	error_handler = logging.FileHandler(ERROR_LOG_FILE, encoding="utf-8")
	error_handler.setLevel(logging.ERROR)
	error_handler.setFormatter(formatter)

	logger.addHandler(exec_handler)
	logger.addHandler(error_handler)
	return logger


def setup_env() -> None:
	"""社内プロキシ環境を利用するための環境変数設定。"""
	os.environ["http_proxy"] = "http://proxy50.adm.toyota.co.jp:15520"
	os.environ["https_proxy"] = "http://proxy50.adm.toyota.co.jp:15520"


class JudgementResult(BaseModel):
	should_append: bool
	reason: str
	confidence: str


class ResponseGenerator:
	"""Azure OpenAI Chat Completions（structured output）用ラッパー。"""

	def __init__(self, logger: Optional[logging.Logger] = None):
		api_base = os.environ["AZURE_OPENAI_ENDPOINT"]
		api_key = os.environ["AZURE_OPENAI_API_KEY"]
		self.deployment_name = os.environ["AZURE_OPENAI_DEPLOYMENT"]
		api_version = "2025-04-01-preview"
		self.logger = logger or logging.getLogger("vsi_tool_input")

		self.client = AzureOpenAI(
			api_key=api_key,
			api_version=api_version,
			azure_endpoint=api_base,
		)

	def create_format_response(
		self,
		messages: List[Dict[str, Any]],
		response_format: Any,
		reasoning_effort: str = "medium",
		retry: int = 1,
	) -> Optional[Any]:
		# API一時エラーを考慮し、指定回数までリトライする。
		last_exc: Optional[Exception] = None
		for _ in range(max(1, retry + 1)):
			try:
				self.logger.info("LLM request start: deployment=%s effort=%s", self.deployment_name, reasoning_effort)
				resp = self.client.beta.chat.completions.parse(
					model=self.deployment_name,
					messages=messages,
					response_format=response_format,
					reasoning_effort=reasoning_effort,
					store=False,
				)
				self.logger.info("LLM request success: deployment=%s", self.deployment_name)
				return resp.choices[0].message.parsed
			except Exception as exc:
				last_exc = exc
				self.logger.exception("LLM request error: deployment=%s", self.deployment_name)
				print(f"{self.deployment_name} API error: {exc}")

		print(f"LLM call failed after retries: {last_exc}")
		return None


@dataclass(frozen=True)
class DetailPair:
	detail_code: str
	detail_name: str


@dataclass
class MissingCandidate:
	row_idx: int
	condition_no: str
	condition_name: str
	condition_basis: str
	major_code: str
	major_name: str
	detail_code: str
	detail_name: str
	detail_col: int


@dataclass(frozen=True)
class StandardDetailRecord:
	major_code: str
	standardized: bool
	detail_code: str
	name_en: str
	name_kana: str


class VSIToolInput:
	"""要求仕様に沿って Excel2ファイルを処理するコア処理。"""

	def __init__(self, rg: ResponseGenerator, logger: Optional[logging.Logger] = None):
		self.rg = rg
		self.logger = logger or logging.getLogger("vsi_tool_input")
		# 過剰呼び出しで止まらないように見える状態を防ぐ。
		self.max_llm_calls = max(1, int(os.getenv("VSI_MAX_LLM_CALLS", "300")))
		self.max_consecutive_blank_rows = max(10, int(os.getenv("VSI_MAX_CONSECUTIVE_BLANK_ROWS", "200")))

	@staticmethod
	def _normalize_text(value: Any) -> str:
		return "" if value is None else str(value).strip()

	@staticmethod
	def _resolve_simple_cell_reference(value: Any, ws: Worksheet) -> Any:
		"""'=Sheet!B2' 形式の単純参照を実値に解決する。"""
		if value is None:
			return None

		text = str(value).strip()
		m = re.match(r"^=\s*(?:'([^']+)'|([^!]+))!\$?([A-Za-z]+)\$?(\d+)\s*$", text)
		if not m:
			return value

		sheet_name = m.group(1) or m.group(2)
		col_letters = m.group(3).upper()
		row_idx = int(m.group(4))

		if sheet_name not in ws.parent.sheetnames:
			return value

		ref_ws = ws.parent[sheet_name]
		col_idx = 0
		for ch in col_letters:
			col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)
		return ref_ws.cell(row=row_idx, column=col_idx).value

	@staticmethod
	def _normalize_major_code(raw: Any) -> str:
		text = "" if raw is None else str(raw).strip()
		# 先頭の '=' は式/表記揺れとして許容し、コード本体のみ評価する。
		text = re.sub(r"^=+", "", text).strip().upper()
		# 大分類コードは英数字3文字のみを有効値とする。
		if re.fullmatch(r"[A-Z0-9]{3}", text):
			return text
		return ""

	@staticmethod
	def _normalize_major_name(raw: Any) -> str:
		# 大分類名称は3文字制限せず、セル文字列全体を保持する。
		return "" if raw is None else str(raw).strip()

	@staticmethod
	def _parse_detail_pair(raw: Any) -> Optional[DetailPair]:
		text = "" if raw is None else str(raw).strip()
		if not text:
			return None

		text = text.replace("：", ":")
		m = re.match(r'^\s*"?([A-Za-z])"?\s*:\s*"?(.+?)"?\s*$', text)
		if not m:
			return None

		detail_code = m.group(1).upper().strip()
		detail_name = m.group(2).strip().strip('"')
		if not detail_name:
			return None
		return DetailPair(detail_code=detail_code, detail_name=detail_name)

	@staticmethod
	def _extract_detail_codes_ordered(raw: Any) -> List[str]:
		text = "" if raw is None else str(raw)
		ordered: List[str] = []
		seen: Set[str] = set()
		for token in re.findall(r"[A-Za-z]", text):
			code = token.upper()
			if code in seen:
				continue
			seen.add(code)
			ordered.append(code)
		return ordered

	@staticmethod
	def _parse_existing_detail_codes(raw: Any) -> Set[str]:
		return set(VSIToolInput._extract_detail_codes_ordered(raw))

	@staticmethod
	def _append_detail_code(existing: Any, code: str) -> str:
		# 既存セルを英字1文字コード列として正規化し、空白や末尾カンマを除去する。
		normalized_codes = VSIToolInput._extract_detail_codes_ordered(existing)
		seen = set(normalized_codes)

		new_code = code.strip().upper()
		if new_code and new_code not in seen:
			normalized_codes.append(new_code)

		return ",".join(normalized_codes)

	def _llm_should_append(
		self,
		condition_name: str,
		condition_basis: str,
		detail_name: str,
		reasoning_effort: str = "medium",
	) -> Tuple[bool, str, str]:
		prompt = f"""## 役割
車両条件の設定基準に対して、候補細目名称が追記対象かを判定してください。

## 判定ルール
- 設定基準に合致すると判断できる場合のみ should_append=true
- 判断根拠が不足する場合は false
- 推測ではなく、与えられた文のみで判断
- 細目名称のW/OはWithoutを意味しており、その大分類仕様が装備されないことを指す
- confidence は以下のみを返す
  - High: 記載方法に従って判断できる場合
  - Low: 記載方法の内容が細目名称からは判断できない場合

## 入力
- 車両条件名称: {condition_name}
- 条件の設定基準: {condition_basis}
- 候補細目名称: {detail_name}

## 出力
- should_append: true/false
- reason: 1文で根拠
- confidence: High/Low
"""
		messages = [{"role": "user", "content": prompt}]
		result: Optional[JudgementResult] = self.rg.create_format_response(
			messages=messages,
			response_format=JudgementResult,
			reasoning_effort=reasoning_effort,
			retry=1,
		)

		if result is None:
			return False, "LLM判定失敗", "Low"
		confidence = "High" if str(result.confidence).strip().lower() == "high" else "Low"
		return result.should_append, result.reason, confidence

	@staticmethod
	def _write_judgement_sheet(wb_vehicle, judgement_rows: List[Dict[str, Any]]) -> None:
		"""LLM判定結果を別シートに書き出す。"""
		sheet_name = "LLM判定結果"
		if sheet_name in wb_vehicle.sheetnames:
			del wb_vehicle[sheet_name]

		ws_judgement = wb_vehicle.create_sheet(title=sheet_name)
		headers = [
			"row",
			"condition_no",
			"condition_name",
			"major_code",
			"major_name",
			"detail_code",
			"detail_name",
			"basis",
			"should_append",
			"reason",
			"confidence",
		]
		ws_judgement.append(headers)

		for item in judgement_rows:
			ws_judgement.append([item.get(col, "") for col in headers])

		# 指定ピクセル幅を openpyxl の列幅（文字数ベース）に変換する。
		def pixels_to_width(px: int) -> float:
			return max(0.0, (px - 5) / 7)

		ws_judgement.column_dimensions["C"].width = pixels_to_width(176)
		ws_judgement.column_dimensions["E"].width = pixels_to_width(241)
		ws_judgement.column_dimensions["G"].width = pixels_to_width(244)
		ws_judgement.column_dimensions["H"].width = pixels_to_width(441)
		ws_judgement.column_dimensions["J"].width = pixels_to_width(703)

		last_row = ws_judgement.max_row
		last_col = ws_judgement.max_column
		wrap_alignment = Alignment(wrap_text=True, vertical="top")
		thin = Side(style="thin", color="000000")
		all_border = Border(left=thin, right=thin, top=thin, bottom=thin)

		for row in ws_judgement.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
			for cell in row:
				cell.alignment = wrap_alignment
				cell.border = all_border

		table_ref = f"A1:K{last_row}"
		table = Table(displayName="LLMJudgementTable", ref=table_ref)
		table.tableStyleInfo = TableStyleInfo(
			name="TableStyleMedium2",
			showFirstColumn=False,
			showLastColumn=False,
			showRowStripes=True,
			showColumnStripes=False,
		)
		ws_judgement.add_table(table)

	@staticmethod
	def _write_standard_check_sheet(wb_vehicle, standard_check_rows: List[Dict[str, Any]]) -> None:
		"""標準大分類/細目一覧との突合結果を別シートに書き出す。"""
		sheet_name = "標準細目チェック"
		if sheet_name in wb_vehicle.sheetnames:
			del wb_vehicle[sheet_name]

		ws_check = wb_vehicle.create_sheet(title=sheet_name)
		headers = [
			"level",
			"category",
			"major_code",
			"detail_code",
			"vsi_detail_name",
			"standard_name_en",
			"standard_name_kana",
			"vehicle_row",
			"vehicle_condition_no",
			"vehicle_condition_name",
			"message",
		]
		ws_check.append(headers)

		for item in standard_check_rows:
			ws_check.append([item.get(col, "") for col in headers])

	def _unmerge_and_fill_for_title_rows(self, ws: Worksheet, start_row: int = 9, end_row: int = 11) -> None:
		target_ranges = []
		for merged_range in list(ws.merged_cells.ranges):
			if merged_range.max_row < start_row or merged_range.min_row > end_row:
				continue
			target_ranges.append(merged_range)

		for merged_range in target_ranges:
			val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
			ws.unmerge_cells(str(merged_range))
			for r in range(merged_range.min_row, merged_range.max_row + 1):
				for c in range(merged_range.min_col, merged_range.max_col + 1):
					ws.cell(row=r, column=c, value=val)

	def _read_country_specs(self, wb_country) -> Tuple[Dict[str, str], Dict[str, Dict[str, str]], List[str]]:
		# STEP1-2: 国別仕様から「大分類コード -> 大分類名称」と「大分類コード -> 細目一覧」を作る。
		ws = wb_country.active
		self._unmerge_and_fill_for_title_rows(ws, 9, 11)

		max_col = ws.max_column
		max_row = ws.max_row
		major_code_to_name: Dict[str, str] = {}
		major_to_details: Dict[str, Dict[str, str]] = {}
		memo: List[str] = []

		for col in range(2, max_col + 1):
			raw_row9 = ws.cell(row=9, column=col).value
			raw_row10 = ws.cell(row=10, column=col).value

			# ファイルにより9/10行目のコード・名称の配置が逆転するため両対応する。
			code_from_row9 = self._normalize_major_code(raw_row9)
			code_from_row10 = self._normalize_major_code(raw_row10)

			if code_from_row9 and not code_from_row10:
				major_code = code_from_row9
				major_name = self._normalize_major_name(raw_row10)
			elif code_from_row10 and not code_from_row9:
				major_code = code_from_row10
				major_name = self._normalize_major_name(raw_row9)
			elif code_from_row9 and code_from_row10:
				# 両方コード判定できる場合は従来想定（9行目コード）を優先。
				major_code = code_from_row9
				major_name = self._normalize_major_name(raw_row10)
			else:
				major_code = ""
				major_name = ""

			if not major_code:
				continue

			if major_code not in major_code_to_name:
				major_code_to_name[major_code] = major_name
			major_to_details.setdefault(major_code, {})

			# 12行目以降は「細目コード:細目名」の定義行。
			blank_streak = 0
			for row in range(12, max_row + 1):
				pair = self._parse_detail_pair(ws.cell(row=row, column=col).value)
				if pair is None:
					blank_streak += 1
					if blank_streak >= self.max_consecutive_blank_rows:
						break
					continue
				blank_streak = 0
				major_to_details[major_code][pair.detail_code] = pair.detail_name

		memo.append(f"大分類数: {len(major_code_to_name)}")
		total_detail = sum(len(v) for v in major_to_details.values())
		memo.append(f"細目ペア総数(重複除去後): {total_detail}")
		return major_code_to_name, major_to_details, memo

	def _get_sheet_by_name_or_first(self, wb, preferred_name: str):
		return wb[preferred_name] if preferred_name in wb.sheetnames else wb.worksheets[0]

	@staticmethod
	def _normalize_name_for_compare(raw: Any) -> str:
		"""英語/カナ差分の表記揺れ吸収用に、空白と記号を除去して大文字化する。"""
		text = "" if raw is None else str(raw).strip()
		text = text.upper()
		text = re.sub(r"[\s\u3000\-_/()\[\]{}:：,，.．]+", "", text)
		return text

	def _read_standard_detail_master(self, wb_standard) -> Tuple[Dict[Tuple[str, str], StandardDetailRecord], List[str]]:
		"""標準大分類/細目一覧（Sheet1）を読み込む。"""
		ws_std = wb_standard["Sheet1"] if "Sheet1" in wb_standard.sheetnames else wb_standard.worksheets[0]
		result: Dict[Tuple[str, str], StandardDetailRecord] = {}
		memo: List[str] = []

		empty_streak = 0
		for row in range(2, ws_std.max_row + 1):
			major_code = self._normalize_major_code(ws_std.cell(row=row, column=2).value)  # B
			detail_code = self._normalize_text(ws_std.cell(row=row, column=6).value).upper()  # F
			if not major_code and not detail_code:
				empty_streak += 1
				if empty_streak >= self.max_consecutive_blank_rows:
					break
				continue
			empty_streak = 0
			if not major_code or not re.fullmatch(r"[A-Z]", detail_code):
				continue

			standardized_text = self._normalize_text(ws_std.cell(row=row, column=5).value)  # E
			standardized = standardized_text == "標準化"

			rec = StandardDetailRecord(
				major_code=major_code,
				standardized=standardized,
				detail_code=detail_code,
				name_en=self._normalize_text(ws_std.cell(row=row, column=7).value),  # G
				name_kana=self._normalize_text(ws_std.cell(row=row, column=9).value),  # I
			)
			result[(major_code, detail_code)] = rec

		std_count = sum(1 for r in result.values() if r.standardized)
		non_std_count = sum(1 for r in result.values() if not r.standardized)
		memo.append(f"標準細目マスタ読込件数: {len(result)}")
		memo.append(f"標準化件数: {std_count}")
		memo.append(f"標準化未実施件数: {non_std_count}")
		return result, memo

	def _collect_standard_check_rows(
		self,
		ws_vehicle: Worksheet,
		major_to_details: Dict[str, Dict[str, str]],
		standard_map: Dict[Tuple[str, str], StandardDetailRecord],
	) -> Tuple[List[Dict[str, Any]], List[str]]:
		"""国別仕様/車両情報と標準マスタを突合し、Warning/Cautionを抽出する。"""
		rows: List[Dict[str, Any]] = []
		memo: List[str] = []

		# Warning: 国別仕様の細目名称と標準マスタ（英語/カナ）の不一致
		warning_count = 0
		for major_code, details in major_to_details.items():
			for detail_code, country_name in details.items():
				rec = standard_map.get((major_code, detail_code))
				if rec is None or not rec.standardized:
					continue

				country_norm = self._normalize_name_for_compare(country_name)
				en_norm = self._normalize_name_for_compare(rec.name_en)
				kana_norm = self._normalize_name_for_compare(rec.name_kana)
				matched_names = {
					en_norm,
					kana_norm,
					f"{en_norm}{kana_norm}" if en_norm or kana_norm else "",
					f"{kana_norm}{en_norm}" if en_norm or kana_norm else "",
				}
				if country_norm and country_norm in matched_names:
					continue

				warning_count += 1
				rows.append(
					{
						"level": "Warning",
						"category": "標準名称不一致",
						"major_code": major_code,
						"detail_code": detail_code,
						"vsi_detail_name": country_name,
						"standard_name_en": rec.name_en,
						"standard_name_kana": rec.name_kana,
						"vehicle_row": "",
						"vehicle_condition_no": "",
						"vehicle_condition_name": "",
						"message": "国別仕様の細目名称が標準マスタ（英語/カナ）と一致しません。",
					}
				)

		# Caution: 車両情報一覧に記載済みだが、標準化未実施の細目
		caution_count = 0
		condition_sets = [4, 7, 10, 13, 16]  # D/G/J/M/P
		for row in range(2, ws_vehicle.max_row + 1):
			condition_name = self._normalize_text(ws_vehicle.cell(row=row, column=2).value)
			if not condition_name:
				break
			condition_no = self._normalize_text(ws_vehicle.cell(row=row, column=1).value)

			for major_col in condition_sets:
				detail_col = major_col + 1
				major_code = self._normalize_major_code(ws_vehicle.cell(row=row, column=major_col).value)
				if not major_code:
					break

				for detail_code in self._extract_detail_codes_ordered(ws_vehicle.cell(row=row, column=detail_col).value):
					rec = standard_map.get((major_code, detail_code))
					if rec is None or rec.standardized:
						continue

					caution_count += 1
					rows.append(
						{
							"level": "Caution",
							"category": "標準化未実施細目",
							"major_code": major_code,
							"detail_code": detail_code,
							"vsi_detail_name": major_to_details.get(major_code, {}).get(detail_code, ""),
							"standard_name_en": rec.name_en,
							"standard_name_kana": rec.name_kana,
							"vehicle_row": row,
							"vehicle_condition_no": condition_no,
							"vehicle_condition_name": condition_name,
							"message": "車両情報一覧に記載されていますが、標準マスタでは標準化未実施です。",
						}
					)

		memo.append(f"Warning件数(標準名称不一致): {warning_count}")
		memo.append(f"Caution件数(標準化未実施細目): {caution_count}")
		return rows, memo

	def _build_basis_map(self, ws_basis: Worksheet) -> Dict[str, str]:
		# 記載方法シートから、No/名称/行番号の3キーで基準文を引ける辞書を作る。
		result: Dict[str, str] = {}
		empty_streak = 0
		for row in range(2, ws_basis.max_row + 1):
			raw_no = self._resolve_simple_cell_reference(ws_basis.cell(row=row, column=1).value, ws_basis)
			raw_name = self._resolve_simple_cell_reference(ws_basis.cell(row=row, column=2).value, ws_basis)
			raw_basis = self._resolve_simple_cell_reference(ws_basis.cell(row=row, column=3).value, ws_basis)

			no = self._normalize_text(raw_no)
			name = self._normalize_text(raw_name)
			basis = self._normalize_text(raw_basis)
			if not no and not name and not basis:
				empty_streak += 1
				if empty_streak >= self.max_consecutive_blank_rows:
					break
				continue
			empty_streak = 0
			if no:
				result[f"NO:{no}"] = basis
			if name and basis and f"NAME:{name}" not in result:
				result[f"NAME:{name}"] = basis
			if basis:
				# 車両情報一覧と記載方法の並びが同じ場合のフォールバックキー
				result[f"ROW:{row}"] = basis
		return result

	def _collect_missing_candidates(
		self,
		ws_vehicle: Worksheet,
		basis_map: Dict[str, str],
		major_code_to_name: Dict[str, str],
		major_to_details: Dict[str, Dict[str, str]],
	) -> Tuple[List[MissingCandidate], List[str]]:
		# STEP3-4: 車両情報一覧を走査し、各大分類に対して不足している細目コード候補を列挙する。
		memo: List[str] = []
		candidates: List[MissingCandidate] = []

		condition_sets = [4, 7, 10, 13, 16]  # D/G/J/M/P

		for row in range(2, ws_vehicle.max_row + 1):
			condition_name = self._normalize_text(ws_vehicle.cell(row=row, column=2).value)
			if not condition_name:
				# 車両条件が空になった時点でデータ終端とみなす。
				break

			condition_no = self._normalize_text(ws_vehicle.cell(row=row, column=1).value)
			condition_basis = basis_map.get(f"NO:{condition_no}", "")
			if not condition_basis:
				condition_basis = basis_map.get(f"NAME:{condition_name}", "")
			if not condition_basis:
				condition_basis = basis_map.get(f"ROW:{row}", "")

			for major_col in condition_sets:
				detail_col = major_col + 1
				raw_major_code = ws_vehicle.cell(row=row, column=major_col).value
				major_code = self._normalize_major_code(raw_major_code)

				if not major_code:
					# その行の大分類セットは左から連続前提。空に当たったら次の行へ。
					break
				if major_code not in major_to_details:
					continue

				current_codes = self._parse_existing_detail_codes(
					ws_vehicle.cell(row=row, column=detail_col).value
				)
				all_details = major_to_details[major_code]
				major_name = major_code_to_name.get(major_code, "")

				for detail_code, detail_name in all_details.items():
					if detail_code in current_codes:
						continue
					candidates.append(
						MissingCandidate(
							row_idx=row,
							condition_no=condition_no,
							condition_name=condition_name,
							condition_basis=condition_basis,
							major_code=major_code,
							major_name=major_name,
							detail_code=detail_code,
							detail_name=detail_name,
							detail_col=detail_col,
						)
					)

		memo.append(f"不足細目候補数: {len(candidates)}")
		return candidates, memo

	def run(
		self,
		country_spec_file: BytesIO,
		vehicle_info_file: BytesIO,
		standard_master_file: Optional[BytesIO] = None,
		reasoning_effort: str = "medium",
	) -> Tuple[bytes, str, List[Dict[str, Any]], Dict[str, Any]]:
		# 全体フロー:
		# 1) 入力2ファイル読込
		# 2) 国別仕様抽出 / 設定基準抽出
		# 3) 不足候補抽出
		# 4) LLM判定と車両情報への反映
		# 5) 出力Excelバイト列と画面表示用データを返却
		memo: List[str] = []
		started_at = time.perf_counter()
		self.logger.info("Run started")

		try:
			wb_country = load_workbook(country_spec_file)
			wb_vehicle = load_workbook(vehicle_info_file)
			wb_standard = load_workbook(standard_master_file) if standard_master_file is not None else None
		except Exception:
			self.logger.exception("Workbook load failed")
			raise

		major_code_to_name, major_to_details, m1 = self._read_country_specs(wb_country)
		memo.extend(m1)

		ws_vehicle = self._get_sheet_by_name_or_first(wb_vehicle, "車両情報一覧")
		ws_basis = self._get_sheet_by_name_or_first(wb_vehicle, "記載方法")
		basis_map = self._build_basis_map(ws_basis)

		missing_candidates, m2 = self._collect_missing_candidates(
			ws_vehicle=ws_vehicle,
			basis_map=basis_map,
			major_code_to_name=major_code_to_name,
			major_to_details=major_to_details,
		)
		memo.extend(m2)

		standard_check_rows: List[Dict[str, Any]] = []
		if wb_standard is not None:
			standard_map, m_std = self._read_standard_detail_master(wb_standard)
			memo.extend(m_std)
			standard_check_rows, m_check = self._collect_standard_check_rows(
				ws_vehicle=ws_vehicle,
				major_to_details=major_to_details,
				standard_map=standard_map,
			)
			memo.extend(m_check)
		else:
			memo.append("標準大分類/細目一覧: 未指定のためチェックをスキップ")

		judgement_rows: List[Dict[str, Any]] = []
		appended_count = 0
		llm_call_count = 0
		llm_cache: Dict[Tuple[str, str, str, str], Tuple[bool, str, str]] = {}

		# STEP5: 候補ごとにLLMで追記要否を判定し、trueのみ反映する。
		for cand in missing_candidates:
			cache_key = (
				cand.condition_name,
				cand.condition_basis,
				cand.detail_name,
				reasoning_effort,
			)
			cached = llm_cache.get(cache_key)
			if cached is not None:
				should_append, reason, confidence = cached
			else:
				if llm_call_count >= self.max_llm_calls:
					should_append = False
					reason = f"LLM呼び出し上限({self.max_llm_calls})到達のため未判定"
					confidence = "Low"
				else:
					should_append, reason, confidence = self._llm_should_append(
						condition_name=cand.condition_name,
						condition_basis=cand.condition_basis,
						detail_name=cand.detail_name,
						reasoning_effort=reasoning_effort,
					)
					llm_call_count += 1
				llm_cache[cache_key] = (should_append, reason, confidence)

			judgement_rows.append(
				{
					"row": cand.row_idx,
					"condition_no": cand.condition_no,
					"condition_name": cand.condition_name,
					"major_code": cand.major_code,
					"major_name": cand.major_name,
					"detail_code": cand.detail_code,
					"detail_name": cand.detail_name,
					"basis": cand.condition_basis,
					"should_append": should_append,
					"reason": reason,
					"confidence": confidence,
				}
			)

			if not should_append:
				continue

			current_value = ws_vehicle.cell(row=cand.row_idx, column=cand.detail_col).value
			updated_text = self._append_detail_code(current_value, cand.detail_code)

			cell = ws_vehicle.cell(row=cand.row_idx, column=cand.detail_col)
			cell.value = updated_text
			# openpyxl 3.1.2 では部分文字装飾が不安定なため、追記が入ったセル全体を赤字にする。
			red_font = copy(cell.font)
			red_font.color = Color(rgb="FFFF0000")
			cell.font = red_font
			appended_count += 1

		memo.append(f"追記件数: {appended_count}")
		memo.append(f"LLM実呼び出し件数: {llm_call_count}")
		memo.append(f"LLMキャッシュ件数: {len(llm_cache)}")

		# 画面表示用に、比較情報を大分類単位で再集約する。
		details_by_major: Dict[str, Dict[str, Any]] = {}
		
		for major_code, details in major_to_details.items():
			major_name = major_code_to_name.get(major_code, "")
			details_by_major[major_code] = {
				"major_name": major_name,
				"country_details": sorted(list(details.keys())),  # 大分類に属する全細目コード
				"country_details_with_names": [
					{"細目コード": code, "細目名": details[code]}
					for code in sorted(list(details.keys()))
				],
				"vehicle_details": set(),  # 車両情報に記載された細目コード
			}
		
		# 車両情報側の細目も同じ大分類軸で収集して、左右比較できる形にする。
		condition_col_to_major_col = {5: 4, 8: 7, 11: 10, 14: 13, 17: 16}  # (detail_col: major_col)
		for row in range(2, ws_vehicle.max_row + 1):
			for detail_col, major_col in condition_col_to_major_col.items():
				major_code = self._normalize_major_code(ws_vehicle.cell(row=row, column=major_col).value)
				if not major_code or major_code not in details_by_major:
					continue
				codes = self._parse_existing_detail_codes(ws_vehicle.cell(row=row, column=detail_col).value)
				details_by_major[major_code]["vehicle_details"].update(codes)
		
		# detail_col: major_col のマッピングを確認
		# D:4, E:5, G:7, H:8, J:10, K:11, M:13, N:14, P:16, Q:17
		# 正しくはdetail_col = major_col + 1
		# E(5)がD(4)の細目、H(8)がG(7)の細目...

		details_info = {"by_major_code": details_by_major}
		details_info["standard_checks"] = standard_check_rows

		# ダウンロードExcelに画面表示と同じ判定結果を別シートで追加する。
		self._write_judgement_sheet(wb_vehicle, judgement_rows)
		self._write_standard_check_sheet(wb_vehicle, standard_check_rows)


		out = BytesIO()
		wb_vehicle.save(out)
		out.seek(0)
		elapsed = time.perf_counter() - started_at
		self.logger.info(
			"Run finished: missing_candidates=%s appended=%s llm_calls=%s elapsed_sec=%.2f",
			len(missing_candidates),
			appended_count,
			llm_call_count,
			elapsed,
		)
		return out.getvalue(), "\n".join(memo), judgement_rows, details_info


def main() -> None:
	logger = setup_logger()
	st.set_page_config(page_title="VSI Tool Input", layout="wide")
	st.title("VSI Tool Input")
	st.caption("国別仕様組み合わせと車両情報一覧を突合し、LLM判定で細目追記を実施")

	# 再描画（ボタン押下、DL押下）でも結果を維持するため、画面状態をsession_stateに保持する。
	if "reasoning_effort" not in st.session_state:
		st.session_state["reasoning_effort"] = "medium"
	if "memo_text" not in st.session_state:
		st.session_state["memo_text"] = ""
	if "judgement_rows" not in st.session_state:
		st.session_state["judgement_rows"] = []
	if "output_bytes" not in st.session_state:
		st.session_state["output_bytes"] = None
	if "details_info" not in st.session_state:
		st.session_state["details_info"] = {}
	if "logger" not in st.session_state:
		st.session_state["logger"] = logger

	# LLMクライアント初期化は1回だけ行う。
	if "tool" not in st.session_state:
		try:
			setup_env()
			rg = ResponseGenerator(st.session_state["logger"])
			st.session_state["tool"] = VSIToolInput(rg, st.session_state["logger"])
		except Exception as exc:
			st.session_state["logger"].exception("初期化エラー")
			st.error(f"初期化エラー: {exc}")
			st.stop()

	with st.sidebar:
		st.header("設定")
		st.session_state["reasoning_effort"] = st.selectbox(
			"reasoning_effort",
			["low", "medium", "high"],
			index=["low", "medium", "high"].index(st.session_state["reasoning_effort"]),
		)

	st.info(
		"1) 国別仕様組み合わせExcelを指定\n"
		"2) 車両情報一覧Excelを指定（車両情報一覧シート・記載方法シートを使用）\n"
		"3) 標準大分類/細目一覧Excelを指定（任意、Sheet1を使用）\n"
		"4) 実行後、追記済みファイルをダウンロード\n"
		f"ログ: 実行={EXEC_LOG_FILE.name} / エラー={ERROR_LOG_FILE.name}"
	)

	country_file = st.file_uploader("国別仕様組み合わせファイル (.xlsx/.xlsm)", type=["xlsx", "xlsm"], key="country")
	vehicle_file = st.file_uploader("車両情報一覧ファイル (.xlsx/.xlsm)", type=["xlsx", "xlsm"], key="vehicle")
	standard_master_file = st.file_uploader("標準大分類/細目一覧ファイル (.xlsx/.xlsm, Sheet1)", type=["xlsx", "xlsm"], key="standard_master")

	button_col1, button_col2 = st.columns([1, 1])
	with button_col1:
		run = st.button("実行", type="primary")
	with button_col2:
		reset = st.button("リセット", type="secondary")

	if reset:
		# 明示的なリセットボタンでのみ、表示・DL対象の結果をクリアする。
		st.session_state["memo_text"] = ""
		st.session_state["judgement_rows"] = []
		st.session_state["output_bytes"] = None
		st.session_state["details_info"] = {}
		st.success("表示結果をリセットしました")

	memo_text = st.session_state["memo_text"]
	judgement_rows = st.session_state["judgement_rows"]
	output_bytes = st.session_state["output_bytes"]
	details_info = st.session_state["details_info"]
	processing_done = False

	if run:
		if country_file is None or vehicle_file is None:
			st.warning("2つのExcelファイルを選択してください。")
		else:
			# 実行時のみ入力ファイルを読み、完了後に結果をsessionへ保存する。
			with st.spinner("処理中です..."):
				try:
					output_bytes, memo_text, judgement_rows, details_info = st.session_state["tool"].run(
						country_spec_file=BytesIO(country_file.read()),
						vehicle_info_file=BytesIO(vehicle_file.read()),
						standard_master_file=BytesIO(standard_master_file.read()) if standard_master_file is not None else None,
						reasoning_effort=st.session_state["reasoning_effort"],
					)
					st.session_state["output_bytes"] = output_bytes
					st.session_state["memo_text"] = memo_text
					st.session_state["judgement_rows"] = judgement_rows
					st.session_state["details_info"] = details_info
					processing_done = True
				except Exception as exc:
					st.session_state["logger"].exception("処理中エラー")
					st.error(f"処理中にエラーが発生しました: {exc}")
			
			if processing_done:
				st.success("✅ 実行完了しました")

	col1, col2 = st.columns([1, 2], gap="large")

	with col1:
		st.subheader("処理メモ")
		st.text_area("memo", value=memo_text, height=250, disabled=True)

		if output_bytes is not None:
			st.download_button(
				label="追記済みExcelをダウンロード",
				data=output_bytes,
				file_name="車両情報一覧_追記済み.xlsx",
				mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			)

	with col2:
		st.subheader("LLM判定結果")
		if judgement_rows:
			st.dataframe(judgement_rows, width='stretch')
			standard_checks = details_info.get("standard_checks", []) if details_info else []
			if standard_checks:
				st.subheader("標準細目チェック結果")
				st.dataframe(standard_checks, width='stretch', hide_index=True)
		else:
			standard_checks = details_info.get("standard_checks", []) if details_info else []
			if standard_checks:
				st.subheader("標準細目チェック結果")
				st.dataframe(standard_checks, width='stretch', hide_index=True)

			if memo_text and "不足細目候補数: 0" in memo_text:
				st.info("📋 不足細目がありません。\n\n国別仕様から抽出した細目が既に車両情報一覧に全て記載されているため、追記対象がありません。")
				
				if details_info and details_info.get("by_major_code"):
					st.write("**大分類ごとの細目比較:**")
					for major_code, info in sorted(details_info["by_major_code"].items()):
						with st.expander(f"{major_code} - {info['major_name']}"):
							col_a, col_b = st.columns(2)
							with col_a:
								st.write("**国別仕様の細目:**")
								st.dataframe(
									info["country_details_with_names"],
									width='stretch',
									hide_index=True,
								)
							with col_b:
								st.write("**車両情報一覧に記載されている細目:**")
								# vehicle_details は set なので list に変換してソート
								vehicle_details_value = info.get("vehicle_details", set())
								if isinstance(vehicle_details_value, set):
									vehicle_codes = sorted(list(vehicle_details_value))
								else:
									vehicle_codes = sorted(list(vehicle_details_value)) if vehicle_details_value else []
								
								if vehicle_codes:
									st.dataframe(
										[{"細目コード": code} for code in vehicle_codes],
										width='stretch',
										hide_index=True,
									)
								else:
									st.write("*なし*")
			elif memo_text:
				st.info("⏳ 結果ここに表示されます（処理中またはフィルタ結果による）。")
			else:
				st.write("📁 Excelファイルを選択して実行してください。")


if __name__ == "__main__":
	main()
